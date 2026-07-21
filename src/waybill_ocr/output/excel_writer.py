import os
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from waybill_ocr.constants import RESULT_WORKBOOK_NAME
from waybill_ocr.container_code.expected_codes import ComparisonReport
from waybill_ocr.error_messages import merge_review_note_with_failure_reason
from waybill_ocr.models import RecognitionResult, RecognitionStatus


HEADERS = ["原始文件名", "识别箱号", "识别状态", "处理耗时ms", "备注"]
INTERNAL_INDEX_SHEET_NAME = "内部索引"
INTERNAL_INDEX_HEADERS = ["原始文件名", "相对路径", "原始识别箱号", "识别来源", "失败原因", "复核候选", "识别状态", "输出相对路径"]
ERROR_ROW_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
COLUMN_WIDTHS = {"A": 52, "B": 18, "C": 14, "D": 14, "E": 52}
COMPARISON_HEADERS = ["预期箱号", "状态", "对应文件", "多余识别箱号", "格式无效"]
MISSING_CODES_FILENAME = "缺失箱号清单.txt"


def write_results(
    results: list[RecognitionResult],
    output_dir: Path,
    workbook_path: Path | None = None,
    comparison_report: ComparisonReport | None = None,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "识别结果"
    sheet.append(HEADERS)
    _apply_result_sheet_widths(sheet)

    for result in results:
        sheet.append(
            [
                result.original_name,
                _display_container_code(result),
                result.status.value,
                result.elapsed_ms,
                _display_review_note(result),
            ]
        )
        if result.status is not RecognitionStatus.SUCCESS:
            _highlight_row(sheet[sheet.max_row][: len(HEADERS)])

    _append_internal_index_sheet(workbook, results)

    if comparison_report is not None:
        _append_comparison_sheet(workbook, comparison_report)

    output_dir.mkdir(parents=True, exist_ok=True)
    if comparison_report is not None:
        _write_missing_codes_file(output_dir, comparison_report.missing_codes)
    target_path = workbook_path or output_dir / RESULT_WORKBOOK_NAME
    _save_workbook_atomically(workbook, target_path)
    return target_path


def _save_workbook_atomically(workbook: Workbook, target_path: Path) -> None:
    temporary_path = target_path.with_name(f".{target_path.stem}.{uuid4().hex}.tmp{target_path.suffix}")
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, target_path)
    finally:
        try:
            temporary_path.unlink()
        except FileNotFoundError:
            pass


def _display_review_note(result: RecognitionResult) -> str:
    if result.status is RecognitionStatus.SUCCESS:
        return result.review_note or ""
    return merge_review_note_with_failure_reason(result.review_note, result.failure_reason) or ""


def _display_container_code(result: RecognitionResult) -> str:
    if result.status is not RecognitionStatus.SUCCESS and result.review_code:
        return result.review_code
    return result.container_code or ""


def _append_internal_index_sheet(workbook, results: list[RecognitionResult]) -> None:
    sheet = workbook.create_sheet(INTERNAL_INDEX_SHEET_NAME)
    sheet.sheet_state = "hidden"
    sheet.append(INTERNAL_INDEX_HEADERS)
    for result in results:
        sheet.append(
            [
                result.original_name,
                result.relative_name or "",
                result.container_code or "",
                result.source.value if result.source else "",
                result.failure_reason or "",
                result.review_code or "",
                result.status.value,
                result.output_relative_path or "",
            ]
        )


def _highlight_row(cells) -> None:
    for cell in cells:
        cell.fill = ERROR_ROW_FILL


def _apply_result_sheet_widths(sheet) -> None:
    for column, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width


def _append_comparison_sheet(workbook, report: ComparisonReport) -> None:
    sheet = workbook.create_sheet("箱号比对")
    sheet.append(COMPARISON_HEADERS)
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 32
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 28
    row_count = max(
        len(report.expected_details),
        len(report.extra_codes),
        len(report.invalid_expected_entries),
    )
    for index in range(row_count):
        detail = _detail_at(report.expected_details, index)
        sheet.append(
            [
                detail.expected_code if detail else "",
                detail.status if detail else "",
                detail.matched_result if detail else "",
                _item_at(report.extra_codes, index),
                _item_at(report.invalid_expected_entries, index),
            ]
        )


def _write_missing_codes_file(output_dir: Path, missing_codes: list[str]) -> None:
    content = "\n".join(missing_codes) + "\n" if missing_codes else "无缺失箱号\n"
    (output_dir / MISSING_CODES_FILENAME).write_text(content, encoding="utf-8")


def _detail_at(items, index: int):
    if index >= len(items):
        return None
    return items[index]


def _item_at(items: list[str], index: int) -> str:
    if index >= len(items):
        return ""
    return items[index]
