import shutil
from collections.abc import Callable
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from time import perf_counter

from openpyxl import load_workbook

from waybill_ocr.cancellation import ProcessingCancelled, raise_if_cancelled
from waybill_ocr.config import AppConfig
from waybill_ocr.constants import RESULT_WORKBOOK_NAME
from waybill_ocr.container_code.expected_codes import compare_expected_codes
from waybill_ocr.container_code.review_candidates import apply_expected_review_code
from waybill_ocr.file_scanner import scan_input_files
from waybill_ocr.models import FileTask, RecognitionResult, RecognitionSource, RecognitionStatus
from waybill_ocr.ocr.base import OcrEngine
from waybill_ocr.output.classifier import copy_result_file
from waybill_ocr.output.excel_writer import INTERNAL_INDEX_SHEET_NAME, write_results
from waybill_ocr.pipeline import process_file
from waybill_ocr.recovery_journal import (
    append_recovery_result,
    clear_recovery_journal,
    load_recovery_results,
    recovery_journal_path,
)

ProgressCallback = Callable[[str], None]


@dataclass(frozen=True)
class ProcessingProgressEvent:
    kind: str
    total: int = 0
    result: RecognitionResult | None = None


@dataclass(frozen=True)
class WorkbookWriteOutcome:
    path: Path | None
    written: bool


ProgressEventCallback = Callable[[ProcessingProgressEvent], None]
RESULT_WRITE_BATCH_SIZE = 5


def process_directory(
    input_dir: Path,
    output_dir: Path,
    config: AppConfig,
    ocr_engine: OcrEngine,
    on_progress: ProgressCallback | None = None,
    cancel_event=None,
    expected_codes: list[str] | None = None,
    expected_invalid_entries: list[str] | None = None,
    on_progress_event: ProgressEventCallback | None = None,
    max_file_workers: int = 1,
) -> list[RecognitionResult]:
    tasks = _exclude_output_tasks(scan_input_files(input_dir), output_dir)
    journal_path = recovery_journal_path(config.state_dir, output_dir) if config.state_dir else None
    existing_results = _load_existing_results(tasks, output_dir, on_progress)
    recovered_results: dict[str, RecognitionResult] = {}
    if journal_path is not None:
        recovered_results = load_recovery_results(journal_path, tasks)
        if recovered_results:
            existing_results.update(recovered_results)
            _emit(on_progress, f"已恢复 {len(recovered_results)} 个尚未写入结果表的处理记录")
    _emit(on_progress, f"\u626b\u63cf\u5230 {len(tasks)} \u4e2a\u6587\u4ef6")
    if tasks:
        _emit(
            on_progress,
            f"\u7ed3\u679c\u8868\u5206\u6279\u5199\u5165\uff1a\u6bcf {RESULT_WRITE_BATCH_SIZE} \u4e2a\u65b0\u7ed3\u679c\u66f4\u65b0\u4e00\u6b21\uff0c"
            "\u4efb\u52a1\u7ed3\u675f\u6216\u53d6\u6d88\u65f6\u4fdd\u5b58\u5269\u4f59\u7ed3\u679c",
        )
    _emit_event(on_progress_event, ProcessingProgressEvent(kind="scanned", total=len(tasks)))

    results: list[RecognitionResult] = []
    workbook_path: Path | None = None
    latest_comparison_report = None
    # Classification stays immediate; workbook rebuilds are amortized in small batches.
    pending_workbook_results = 0
    new_result_count = 0
    last_written_result_count = 0
    total = len(tasks)
    skipped_existing_result = False
    recovered_unpersisted_result = bool(recovered_results)
    worker_limit = max(1, min(max_file_workers, 2))
    pending_tasks: list[tuple[int, FileTask]] = []
    for index, task in enumerate(tasks, start=1):
        existing_result = existing_results.get(task.relative_name) or existing_results.get(task.source_path.name)
        if existing_result is None or not _should_skip_existing_result(existing_result, output_dir):
            pending_tasks.append((index, task))
    recognition_executor = None
    recognition_futures = {}
    if worker_limit > 1 and len(pending_tasks) > 1:
        recognition_executor = ThreadPoolExecutor(max_workers=worker_limit)
        _emit(
            on_progress,
            f"\u5e76\u884c\u8bc6\u522b\u5df2\u542f\u7528\uff1a\u540c\u65f6\u5904\u7406 {worker_limit} \u4e2a\u6587\u4ef6\uff0c"
            "\u7ed3\u679c\u4ecd\u6309\u626b\u63cf\u987a\u5e8f\u6c47\u603b",
        )
        recognition_futures = {
            task.relative_name: recognition_executor.submit(
                _process_file_with_progress,
                task,
                index,
                total,
                config,
                ocr_engine,
                on_progress,
                cancel_event,
            )
            for index, task in pending_tasks
        }

    try:
        for index, task in enumerate(tasks, start=1):
            raise_if_cancelled(cancel_event)
            existing_result = existing_results.get(task.relative_name) or existing_results.get(task.source_path.name)
            if existing_result is not None and _should_skip_existing_result(existing_result, output_dir):
                skipped_existing_result = True
                results.append(existing_result)
                _emit(on_progress, f"\u5df2\u8df3\u8fc7\u5df2\u5904\u7406\u6587\u4ef6: {task.relative_name}")
                _emit(on_progress, _format_result_message(existing_result))
                _emit_event(on_progress_event, ProcessingProgressEvent(kind="result", result=existing_result))
                continue

            started_at = perf_counter()
            future = recognition_futures.get(task.relative_name)
            if future is None:
                _emit(on_progress, f"\u5904\u7406\u4e2d: {index}/{total} {task.relative_name}")

            try:
                result = (
                    future.result()
                    if future is not None
                    else process_file(task, config, ocr_engine, cancel_event=cancel_event)
                )
                if result.relative_name is None:
                    result = replace(result, relative_name=task.relative_name)
                result = apply_expected_review_code(result, expected_codes)
                raise_if_cancelled(cancel_event)
                copied_path = _copy_result_file(
                    result,
                    output_dir,
                    existing_result.output_relative_path if existing_result is not None else None,
                )
                result = replace(result, output_relative_path=_output_relative_path(copied_path, output_dir))
            except ProcessingCancelled:
                raise
            except Exception as exc:
                result = _failed_result(task, f"PROCESS_FAILED: {exc}", started_at)
                _emit(on_progress, f"\u6587\u4ef6\u5904\u7406\u5931\u8d25\uff0c\u5df2\u8df3\u8fc7: {task.relative_name} ({exc})")
                result = _copy_failed_result(
                    result,
                    output_dir,
                    on_progress,
                    previous_output_relative_path=(
                        existing_result.output_relative_path if existing_result is not None else None
                    ),
                )

            _append_recovery_result(journal_path, result, on_progress)
            results.append(result)
            new_result_count += 1
            pending_workbook_results += 1
            _emit_event(on_progress_event, ProcessingProgressEvent(kind="result", result=result))
            if pending_workbook_results >= RESULT_WRITE_BATCH_SIZE:
                latest_comparison_report = _comparison_report(
                    expected_codes,
                    results,
                    expected_invalid_entries,
                )
                write_outcome = _write_results(
                    results,
                    output_dir,
                    workbook_path,
                    on_progress,
                    latest_comparison_report,
                )
                workbook_path = write_outcome.path
                if write_outcome.written:
                    last_written_result_count = len(results)
                    _clear_recovery_if_persisted(journal_path, write_outcome, output_dir)
                pending_workbook_results = 0
            _emit(on_progress, _format_result_message(result))

        if latest_comparison_report is None or last_written_result_count != len(results):
            latest_comparison_report = _comparison_report(expected_codes, results, expected_invalid_entries)
        should_write_final_results = bool(results) and (
            pending_workbook_results > 0
            or (new_result_count > 0 and last_written_result_count != len(results))
            or recovered_unpersisted_result
            or (
                skipped_existing_result
                and workbook_path is None
                and latest_comparison_report is not None
            )
        )
        if should_write_final_results:
            write_outcome = _write_results(
                results,
                output_dir,
                workbook_path,
                on_progress,
                latest_comparison_report,
            )
            workbook_path = write_outcome.path
            _clear_recovery_if_persisted(journal_path, write_outcome, output_dir)
        _emit_comparison_report(latest_comparison_report, on_progress)
        _emit(on_progress, "\u5904\u7406\u5b8c\u6210")
        _emit_event(on_progress_event, ProcessingProgressEvent(kind="complete"))
        return results
    except ProcessingCancelled:
        if results:
            write_outcome = _write_results(
                results,
                output_dir,
                workbook_path,
                on_progress,
                _comparison_report(expected_codes, results, expected_invalid_entries),
            )
            _clear_recovery_if_persisted(journal_path, write_outcome, output_dir)
        _emit(on_progress, f"\u5df2\u53d6\u6d88\uff1a\u5df2\u5904\u7406 {len(results)}/{total}")
        _emit_event(on_progress_event, ProcessingProgressEvent(kind="cancelled"))
        return results
    finally:
        if recognition_executor is not None:
            for future in recognition_futures.values():
                future.cancel()
            recognition_executor.shutdown(wait=True, cancel_futures=True)
        _cleanup_work_dir(config)


def _process_file_with_progress(
    task: FileTask,
    index: int,
    total: int,
    config: AppConfig,
    ocr_engine: OcrEngine,
    on_progress: ProgressCallback | None,
    cancel_event,
) -> RecognitionResult:
    raise_if_cancelled(cancel_event)
    _emit(on_progress, f"\u5904\u7406\u4e2d: {index}/{total} {task.relative_name}")
    return process_file(task, config, ocr_engine, cancel_event=cancel_event)


def _should_skip_existing_result(result: RecognitionResult, output_dir: Path) -> bool:
    if result.status is not RecognitionStatus.SUCCESS:
        return False
    if not result.output_relative_path:
        return True
    recorded_path = _recorded_output_path(output_dir, result.output_relative_path)
    return recorded_path is not None and recorded_path.is_file()


def _failed_result(task: FileTask, reason: str, started_at: float) -> RecognitionResult:
    elapsed_ms = int((perf_counter() - started_at) * 1000)
    return RecognitionResult(
        source_path=task.source_path,
        original_name=task.source_path.name,
        status=RecognitionStatus.UNRECOGNIZED,
        container_code=None,
        source=None,
        failure_reason=reason,
        ocr_text="",
        elapsed_ms=elapsed_ms,
        relative_name=task.relative_name,
    )


def _copy_failed_result(
    result: RecognitionResult,
    output_dir: Path,
    on_progress: ProgressCallback | None,
    previous_output_relative_path: str | None = None,
) -> RecognitionResult:
    try:
        copied_path = _copy_result_file(result, output_dir, previous_output_relative_path)

        return replace(result, output_relative_path=_output_relative_path(copied_path, output_dir))
    except Exception as exc:
        _emit(on_progress, f"\u5931\u8d25\u6587\u4ef6\u590d\u5236\u5230\u672a\u8bc6\u522b\u76ee\u5f55\u5931\u8d25\uff0c\u5df2\u7ee7\u7eed: {result.original_name} ({exc})")
        return result


def _append_recovery_result(
    journal_path: Path | None,
    result: RecognitionResult,
    on_progress: ProgressCallback | None,
) -> None:
    if journal_path is None:
        return
    try:
        append_recovery_result(journal_path, result)
    except OSError as exc:
        _emit(on_progress, f"处理状态保存失败，程序将继续处理: {exc}")


def _clear_recovery_if_persisted(
    journal_path: Path | None,
    outcome: WorkbookWriteOutcome,
    output_dir: Path,
) -> None:
    if journal_path is None or not outcome.written or outcome.path is None:
        return
    if outcome.path.resolve() == (output_dir / RESULT_WORKBOOK_NAME).resolve():
        clear_recovery_journal(journal_path)


def _output_relative_path(output_path: Path, output_dir: Path) -> str:
    return output_path.resolve().relative_to(output_dir.resolve()).as_posix()


def _copy_result_file(
    result: RecognitionResult,
    output_dir: Path,
    previous_output_relative_path: str | None,
) -> Path:
    if previous_output_relative_path is None:
        return copy_result_file(result, output_dir)
    return copy_result_file(
        result,
        output_dir,
        previous_output_relative_path=previous_output_relative_path,
    )


def _recorded_output_path(output_dir: Path, relative_path: str) -> Path | None:
    relative = Path(relative_path)
    if relative.is_absolute():
        return None
    output_root = output_dir.resolve()
    candidate = (output_root / relative).resolve()
    try:
        candidate.relative_to(output_root)
    except ValueError:
        return None
    return candidate


def _write_results(
    results: list[RecognitionResult],
    output_dir: Path,
    workbook_path: Path | None,
    on_progress: ProgressCallback | None,
    comparison_report=None,
) -> WorkbookWriteOutcome:
    try:
        written_path = write_results(
            results,
            output_dir,
            workbook_path=workbook_path,
            comparison_report=comparison_report,
        )
        return WorkbookWriteOutcome(path=written_path, written=True)
    except PermissionError as exc:
        if workbook_path is not None:
            _emit(on_progress, f"\u7ed3\u679c\u8868\u5199\u5165\u5931\u8d25\uff0c\u5df2\u7ee7\u7eed\u5904\u7406: {exc}")
            return WorkbookWriteOutcome(path=workbook_path, written=False)

        fallback_path = _backup_workbook_path(output_dir)
        _emit(on_progress, f"{RESULT_WORKBOOK_NAME} \u88ab\u5360\u7528\uff0c\u5df2\u6539\u5199\u5907\u7528\u7ed3\u679c\u8868: {fallback_path.name}")
        try:
            written_path = write_results(
                results,
                output_dir,
                workbook_path=fallback_path,
                comparison_report=comparison_report,
            )
            return WorkbookWriteOutcome(path=written_path, written=True)
        except PermissionError as fallback_exc:
            _emit(on_progress, f"\u7ed3\u679c\u8868\u5199\u5165\u5931\u8d25\uff0c\u5df2\u7ee7\u7eed\u5904\u7406: {fallback_exc}")
            return WorkbookWriteOutcome(path=None, written=False)


def _load_existing_results(
    tasks: list[FileTask],
    output_dir: Path,
    on_progress: ProgressCallback | None,
) -> dict[str, RecognitionResult]:
    workbook_path = output_dir / RESULT_WORKBOOK_NAME
    if not workbook_path.exists():
        return {}

    tasks_by_relative = {task.relative_name: task for task in tasks}
    tasks_by_unique_name = _tasks_by_unique_name(tasks)
    workbook = None
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook[INTERNAL_INDEX_SHEET_NAME] if INTERNAL_INDEX_SHEET_NAME in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = {name: index for index, name in enumerate(next(rows, [])) if name}
        results: dict[str, RecognitionResult] = {}
        for row in rows:
            original_name = _optional_string(_row_value(row, headers, "原始文件名"))
            relative_name = _optional_string(_row_value(row, headers, "相对路径"))
            task = _task_for_existing_row(relative_name, original_name, tasks_by_relative, tasks_by_unique_name)
            if task is None or not original_name:
                continue

            status = _recognition_status(_row_value(row, headers, "识别状态"))
            if status is None:
                continue

            result_relative_name = relative_name or task.relative_name
            results[result_relative_name] = RecognitionResult(
                source_path=task.source_path,
                original_name=original_name,
                status=status,
                container_code=(
                    _optional_string(_row_value(row, headers, "原始识别箱号"))
                    or _optional_string(_row_value(row, headers, "识别箱号"))
                ),
                source=_recognition_source(_row_value(row, headers, "识别来源")),
                failure_reason=_optional_string(_row_value(row, headers, "失败原因")),
                ocr_text="",
                elapsed_ms=_optional_int(_row_value(row, headers, "处理耗时ms")),
                relative_name=result_relative_name,
                review_note=_optional_string(_row_value(row, headers, "备注")),
                review_code=_optional_string(_row_value(row, headers, "复核候选")),
                output_relative_path=_optional_string(_row_value(row, headers, "输出相对路径")),
            )
        return results
    except Exception as exc:
        _emit(on_progress, f"历史结果读取失败，已重新处理全部文件: {exc}")
        return {}
    finally:
        if workbook is not None:
            workbook.close()

def _tasks_by_unique_name(tasks: list[FileTask]) -> dict[str, FileTask]:
    counts: dict[str, int] = {}
    for task in tasks:
        counts[task.source_path.name] = counts.get(task.source_path.name, 0) + 1
    return {task.source_path.name: task for task in tasks if counts[task.source_path.name] == 1}


def _task_for_existing_row(
    relative_name: str | None,
    original_name: str | None,
    tasks_by_relative: dict[str, FileTask],
    tasks_by_unique_name: dict[str, FileTask],
) -> FileTask | None:
    if relative_name and relative_name in tasks_by_relative:
        return tasks_by_relative[relative_name]
    if original_name:
        return tasks_by_unique_name.get(original_name)
    return None

def _row_value(row, headers: dict[str, int], name: str):
    index = headers.get(name)
    if index is None or index >= len(row):
        return None
    return row[index]


def _recognition_status(value) -> RecognitionStatus | None:
    if value is None:
        return None
    try:
        return RecognitionStatus(str(value))
    except ValueError:
        return None


def _recognition_source(value) -> RecognitionSource | None:
    if value is None or value == "":
        return None
    try:
        return RecognitionSource(str(value))
    except ValueError:
        return None


def _optional_string(value) -> str | None:
    if value is None or value == "":
        return None
    return str(value)


def _optional_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _comparison_report(
    expected_codes: list[str] | None,
    results: list[RecognitionResult],
    expected_invalid_entries: list[str] | None = None,
):
    if expected_codes is None and not expected_invalid_entries:
        return None
    return compare_expected_codes(expected_codes or [], results, expected_invalid_entries)


def _emit_comparison_report(
    report,
    on_progress: ProgressCallback | None,
) -> None:
    if report is None:
        return

    _emit(
        on_progress,
        f"\u7bb1\u53f7\u6bd4\u5bf9: \u5df2\u5339\u914d {len(report.matched_codes)}, \u7f3a\u5931 {len(report.missing_codes)}, "
        f"\u591a\u4f59 {len(report.extra_codes)}, \u683c\u5f0f\u65e0\u6548 {len(report.invalid_expected_entries)}",
    )
    if report.missing_codes:
        _emit(on_progress, f"\u7f3a\u5931\u7bb1\u53f7: {', '.join(report.missing_codes)}")
    if report.invalid_expected_entries:
        _emit(on_progress, f"\u683c\u5f0f\u65e0\u6548\u6e05\u5355\u9879: {', '.join(report.invalid_expected_entries)}")


def _backup_workbook_path(output_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return output_dir / f"\u8bc6\u522b\u7ed3\u679c-\u5907\u4efd-{timestamp}.xlsx"


def _exclude_output_tasks(tasks: list[FileTask], output_dir: Path) -> list[FileTask]:
    resolved_output = output_dir.resolve()
    filtered_tasks = []
    for task in tasks:
        resolved_source = task.source_path.resolve()
        if resolved_source == resolved_output or resolved_output in resolved_source.parents:
            continue
        filtered_tasks.append(task)
    return filtered_tasks


def _format_result_message(result: RecognitionResult) -> str:
    detail = result.container_code or result.failure_reason or "\u65e0\u7bb1\u53f7"
    return f"\u7ed3\u679c: {result.original_name} -> {result.status.value} ({detail})"


def _cleanup_work_dir(config: AppConfig) -> None:
    if config.work_dir:
        shutil.rmtree(config.work_dir, ignore_errors=True)


def _close_iterator(iterator) -> None:
    close = getattr(iterator, "close", None)
    if close:
        close()


def _emit_event(on_progress_event: ProgressEventCallback | None, event: ProcessingProgressEvent) -> None:
    if on_progress_event:
        on_progress_event(event)


def _emit(on_progress: ProgressCallback | None, message: str) -> None:
    if on_progress:
        on_progress(message)
