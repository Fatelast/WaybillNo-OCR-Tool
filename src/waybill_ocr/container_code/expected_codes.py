import csv
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from waybill_ocr.container_code.extractor import extract_candidates
from waybill_ocr.models import RecognitionResult, RecognitionStatus

EXPECTED_STATUS_RECOGNIZED = "已识别"
EXPECTED_STATUS_REVIEW = "待确认命中"
EXPECTED_STATUS_MISSING = "缺失"


@dataclass(frozen=True)
class ExpectedCodeInspection:
    valid_codes: list[str]
    duplicate_codes: list[str]
    invalid_entries: list[str]

    @property
    def valid_count(self) -> int:
        return len(self.valid_codes)

    @property
    def duplicate_count(self) -> int:
        return len(self.duplicate_codes)

    @property
    def invalid_count(self) -> int:
        return len(self.invalid_entries)


@dataclass(frozen=True)
class ExpectedCodeDetail:
    expected_code: str
    status: str
    matched_result: str = ""


@dataclass(frozen=True)
class ComparisonReport:
    expected_codes: list[str]
    recognized_codes: list[str]
    matched_codes: list[str]
    missing_codes: list[str]
    extra_codes: list[str]
    invalid_expected_entries: list[str] = field(default_factory=list)
    expected_details: list[ExpectedCodeDetail] = field(default_factory=list)


def read_expected_codes(path: Path) -> list[str]:
    return inspect_expected_codes(path).valid_codes


def inspect_expected_codes(path: Path) -> ExpectedCodeInspection:
    entries = _read_expected_entries(path)
    valid_codes: list[str] = []
    duplicate_codes: list[str] = []
    invalid_entries: list[str] = []

    for entry in entries:
        normalized_entry = entry.strip()
        if not normalized_entry:
            continue

        candidates = extract_candidates(normalized_entry)
        if not candidates:
            invalid_entries.append(normalized_entry)
            continue

        for code in candidates:
            if code in valid_codes:
                if code not in duplicate_codes:
                    duplicate_codes.append(code)
                continue
            valid_codes.append(code)

    return ExpectedCodeInspection(
        valid_codes=valid_codes,
        duplicate_codes=duplicate_codes,
        invalid_entries=invalid_entries,
    )


def compare_expected_codes(
    expected_codes: list[str],
    results: list[RecognitionResult],
    invalid_expected_entries: list[str] | None = None,
) -> ComparisonReport:
    normalized_expected = _dedupe(expected_codes)
    recognized_codes = _recognized_success_codes(results)
    recognized_set = set(recognized_codes)
    expected_set = set(normalized_expected)
    review_hits = _review_hits_by_code(results)
    success_hits = _success_hits_by_code(results)

    matched_codes = [code for code in normalized_expected if code in recognized_set]
    missing_codes = [code for code in normalized_expected if code not in recognized_set]
    extra_codes = [code for code in recognized_codes if code not in expected_set]
    expected_details = [
        _expected_detail(code, success_hits, review_hits) for code in normalized_expected
    ]
    return ComparisonReport(
        expected_codes=normalized_expected,
        recognized_codes=recognized_codes,
        matched_codes=matched_codes,
        missing_codes=missing_codes,
        extra_codes=extra_codes,
        invalid_expected_entries=invalid_expected_entries or [],
        expected_details=expected_details,
    )


def _expected_detail(
    code: str,
    success_hits: dict[str, RecognitionResult],
    review_hits: dict[str, RecognitionResult],
) -> ExpectedCodeDetail:
    success_result = success_hits.get(code)
    if success_result is not None:
        return ExpectedCodeDetail(code, EXPECTED_STATUS_RECOGNIZED, success_result.original_name)
    review_result = review_hits.get(code)
    if review_result is not None:
        return ExpectedCodeDetail(code, EXPECTED_STATUS_REVIEW, review_result.original_name)
    return ExpectedCodeDetail(code, EXPECTED_STATUS_MISSING)


def _read_expected_entries(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_expected_entries_from_workbook(path)
    if suffix == ".csv":
        return _read_expected_entries_from_csv(path)
    return _read_expected_entries_from_text(path)


def _read_expected_entries_from_workbook(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        values = []
        for row in workbook.active.iter_rows(values_only=True):
            for value in row:
                if value is not None:
                    values.append(str(value))
        return values
    finally:
        workbook.close()


def _read_expected_entries_from_csv(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8-sig", errors="ignore")
    entries: list[str] = []
    for row in csv.reader(text.splitlines()):
        for value in row:
            if value.strip():
                entries.append(value)
    return entries


def _read_expected_entries_from_text(path: Path) -> list[str]:
    return [line for line in path.read_text(encoding="utf-8-sig", errors="ignore").splitlines()]


def _recognized_success_codes(results: list[RecognitionResult]) -> list[str]:
    codes: list[str] = []
    for result in results:
        if result.status == RecognitionStatus.SUCCESS and result.container_code and result.container_code not in codes:
            codes.append(result.container_code)
    return codes


def _success_hits_by_code(results: list[RecognitionResult]) -> dict[str, RecognitionResult]:
    hits: dict[str, RecognitionResult] = {}
    for result in results:
        if result.status == RecognitionStatus.SUCCESS and result.container_code and result.container_code not in hits:
            hits[result.container_code] = result
    return hits


def _review_hits_by_code(results: list[RecognitionResult]) -> dict[str, RecognitionResult]:
    hits: dict[str, RecognitionResult] = {}
    for result in results:
        if result.status != RecognitionStatus.SUCCESS and result.review_code and result.review_code not in hits:
            hits[result.review_code] = result
    return hits


def _dedupe(codes: list[str]) -> list[str]:
    unique_codes: list[str] = []
    for code in codes:
        normalized = code.strip().upper()
        if normalized and normalized not in unique_codes:
            unique_codes.append(normalized)
    return unique_codes
