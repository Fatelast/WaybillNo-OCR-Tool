from pathlib import Path

from openpyxl import load_workbook

import waybill_ocr.batch_processor as batch_module
from waybill_ocr.config import AppConfig
from waybill_ocr.container_code.duplicate_results import (
    DUPLICATE_CONTAINER_CODE_REASON,
    mark_duplicate_container_results,
)
from waybill_ocr.models import FileTask, RecognitionResult, RecognitionSource, RecognitionStatus


class FakeOcrEngine:
    pass


def _success_result(source_path: Path) -> RecognitionResult:
    return RecognitionResult(
        source_path=source_path,
        original_name=source_path.name,
        status=RecognitionStatus.SUCCESS,
        container_code="HNKU6331795",
        source=RecognitionSource.OCR,
        failure_reason=None,
        ocr_text="HNKU6331795",
        elapsed_ms=1,
        relative_name=source_path.name,
    )


def test_mark_duplicate_container_results_demotes_every_matching_success(tmp_path: Path):
    results = [
        _success_result(tmp_path / "first.jpg"),
        _success_result(tmp_path / "second.jpg"),
    ]

    updated, changed_indices = mark_duplicate_container_results(results)

    assert changed_indices == (0, 1)
    assert all(result.status is RecognitionStatus.INVALID for result in updated)
    assert all(result.failure_reason == DUPLICATE_CONTAINER_CODE_REASON for result in updated)
    assert all(result.review_code == "HNKU6331795" for result in updated)
    assert all("2 \u4e2a\u6587\u4ef6" in (result.review_note or "") for result in updated)


def test_process_directory_reclassifies_duplicate_success_files(tmp_path: Path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    for name in ("first.jpg", "second.jpg"):
        (input_dir / name).write_bytes(name.encode("ascii"))

    def fake_process_file(task: FileTask, _config, _engine, cancel_event=None):
        return _success_result(task.source_path)

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)
    events = []
    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
        on_progress_event=events.append,
    )

    assert [result.status for result in results] == [RecognitionStatus.INVALID, RecognitionStatus.INVALID]
    assert not list((output_dir / "\u6b63\u786e\u8bc6\u522b").glob("*.jpg"))
    invalid_names = {path.name for path in (output_dir / "\u7bb1\u53f7\u9519\u8bef").glob("*.jpg")}
    assert invalid_names == {
        "HNKU6331795-\u5f85\u786e\u8ba4.jpg",
        "HNKU6331795-\u5f85\u786e\u8ba4-1.jpg",
    }
    assert [event.kind for event in events].count("reclassified") == 1
    assert batch_module.count_retryable_results(output_dir) == 2

    workbook = load_workbook(output_dir / "\u8bc6\u522b\u7ed3\u679c.xlsx", data_only=True)
    try:
        assert [workbook.active.cell(row, 3).value for row in (2, 3)] == [
            RecognitionStatus.INVALID.value,
            RecognitionStatus.INVALID.value,
        ]
        index_sheet = workbook["\u5185\u90e8\u7d22\u5f15"]
        assert [index_sheet.cell(row, 5).value for row in (2, 3)] == [
            DUPLICATE_CONTAINER_CODE_REASON,
            DUPLICATE_CONTAINER_CODE_REASON,
        ]
    finally:
        workbook.close()


def test_count_retryable_results_returns_zero_without_workbook(tmp_path: Path):
    assert batch_module.count_retryable_results(tmp_path) == 0