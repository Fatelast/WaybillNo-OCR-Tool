from pathlib import Path

from openpyxl import load_workbook

import waybill_ocr.batch_processor as batch_module
from waybill_ocr.config import AppConfig
from waybill_ocr.models import FileTask, RecognitionResult, RecognitionSource, RecognitionStatus


class FakeOcrEngine:
    pass


def test_process_directory_classifies_files_and_writes_workbook(tmp_path: Path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    source_path = input_dir / "waybill.jpg"
    source_path.write_bytes(b"fake")
    progress_messages = []

    def fake_process_file(
        task: FileTask,
        _config: AppConfig,
        _ocr_engine: FakeOcrEngine,
        cancel_event=None,
    ) -> RecognitionResult:
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="HNKU6331795",
            elapsed_ms=1,
        )

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)

    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
        on_progress=progress_messages.append,
    )

    assert [result.container_code for result in results] == ["HNKU6331795"]
    assert (output_dir / "正确识别" / "HNKU6331795.jpg").read_bytes() == b"fake"
    workbook = load_workbook(output_dir / "识别结果.xlsx")
    assert workbook.active["B2"].value == "HNKU6331795"
    assert progress_messages == [
        "扫描到 1 个文件",
        "处理中: 1/1 waybill.jpg",
        "结果: waybill.jpg -> 正确识别 (HNKU6331795)",
        "处理完成",
    ]


def test_process_directory_excludes_output_directory_inside_input(tmp_path: Path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = input_dir / "output"
    output_success = output_dir / "正确识别"
    output_success.mkdir(parents=True)
    source_path = input_dir / "waybill.jpg"
    stale_output = output_success / "old.jpg"
    source_path.write_bytes(b"source")
    stale_output.write_bytes(b"old")
    processed = []

    def fake_process_file(
        task: FileTask,
        _config: AppConfig,
        _ocr_engine: FakeOcrEngine,
        cancel_event=None,
    ) -> RecognitionResult:
        processed.append(task.relative_name)
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="HNKU6331795",
            elapsed_ms=1,
        )

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)

    batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
    )

    assert processed == ["waybill.jpg"]


def test_process_directory_cleans_configured_work_dir(tmp_path: Path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    work_dir = tmp_path / "work"
    input_dir.mkdir()
    source_path = input_dir / "waybill.jpg"
    source_path.write_bytes(b"fake")

    def fake_process_file(
        task: FileTask,
        config: AppConfig,
        _ocr_engine: FakeOcrEngine,
        cancel_event=None,
    ) -> RecognitionResult:
        assert config.work_dir == work_dir
        config.work_dir.mkdir(parents=True, exist_ok=True)
        (config.work_dir / "temp.txt").write_text("temp", encoding="utf-8")
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="HNKU6331795",
            elapsed_ms=1,
        )

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)

    batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(work_dir=work_dir),
        ocr_engine=FakeOcrEngine(),
    )

    assert not work_dir.exists()

def test_process_directory_writes_workbook_incrementally_when_later_copy_fails(tmp_path: Path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    first_path = input_dir / "first.jpg"
    second_path = input_dir / "second.jpg"
    first_path.write_bytes(b"first")
    second_path.write_bytes(b"second")

    def fake_process_file(
        task: FileTask,
        _config: AppConfig,
        _ocr_engine: FakeOcrEngine,
        cancel_event=None,
    ) -> RecognitionResult:
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="HNKU6331795",
            elapsed_ms=1,
        )

    def fake_copy_result_file(result: RecognitionResult, _output_dir: Path) -> Path:
        if result.original_name == "second.jpg":
            raise RuntimeError("copy failed")
        return output_dir / "正确识别" / "HNKU6331795.jpg"

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)
    monkeypatch.setattr(batch_module, "copy_result_file", fake_copy_result_file)

    try:
        batch_module.process_directory(
            input_dir=input_dir,
            output_dir=output_dir,
            config=AppConfig(),
            ocr_engine=FakeOcrEngine(),
        )
    except RuntimeError:
        pass

    workbook = load_workbook(output_dir / "识别结果.xlsx")
    assert workbook.active["A2"].value == "first.jpg"
    assert workbook.active["B2"].value == "HNKU6331795"


def test_process_directory_stops_between_files_when_cancelled(tmp_path: Path, monkeypatch):
    import threading

    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    first_path = input_dir / "first.jpg"
    second_path = input_dir / "second.jpg"
    first_path.write_bytes(b"first")
    second_path.write_bytes(b"second")
    cancel_event = threading.Event()
    processed = []
    progress_messages = []

    def fake_process_file(task: FileTask, _config: AppConfig, _ocr_engine: FakeOcrEngine, cancel_event=None) -> RecognitionResult:
        processed.append(task.relative_name)
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="HNKU6331795",
            elapsed_ms=1,
        )

    def fake_copy_result_file(result: RecognitionResult, _output_dir: Path) -> Path:
        cancel_event.set()
        return _output_dir / "正确识别" / result.original_name

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)
    monkeypatch.setattr(batch_module, "copy_result_file", fake_copy_result_file)

    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
        on_progress=progress_messages.append,
        cancel_event=cancel_event,
    )

    assert [result.original_name for result in results] == ["first.jpg"]
    assert processed == ["first.jpg"]
    assert any(message == "已取消：已处理 1/2" for message in progress_messages)



def test_process_directory_writes_backup_workbook_when_default_workbook_is_locked(tmp_path: Path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    source_path = input_dir / "waybill.jpg"
    source_path.write_bytes(b"fake")
    progress_messages = []
    workbook_paths = []

    def fake_process_file(
        task: FileTask,
        _config: AppConfig,
        _ocr_engine: FakeOcrEngine,
        cancel_event=None,
    ) -> RecognitionResult:
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="HNKU6331795",
            elapsed_ms=1,
        )

    def fake_write_results(results, output_dir: Path, workbook_path: Path | None = None, comparison_report=None) -> Path:
        workbook_paths.append(workbook_path)
        if workbook_path is None:
            raise PermissionError("workbook is open")
        assert workbook_path.parent == output_dir
        assert workbook_path.name.startswith("识别结果-备份-")
        assert workbook_path.suffix == ".xlsx"
        return workbook_path

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)
    monkeypatch.setattr(batch_module, "write_results", fake_write_results)

    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
        on_progress=progress_messages.append,
    )

    assert [result.original_name for result in results] == ["waybill.jpg"]
    assert workbook_paths[0] is None
    assert workbook_paths[1] is not None
    assert any("识别结果.xlsx 被占用" in message for message in progress_messages)


def test_process_directory_records_failed_file_and_continues(tmp_path: Path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    first_path = input_dir / "first.jpg"
    second_path = input_dir / "second.jpg"
    first_path.write_bytes(b"first")
    second_path.write_bytes(b"second")
    processed = []
    progress_messages = []

    def fake_process_file(
        task: FileTask,
        _config: AppConfig,
        _ocr_engine: FakeOcrEngine,
        cancel_event=None,
    ) -> RecognitionResult:
        processed.append(task.relative_name)
        if task.relative_name == "first.jpg":
            raise RuntimeError("ocr crashed")
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="HNKU6331795",
            elapsed_ms=1,
        )

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)

    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
        on_progress=progress_messages.append,
    )

    assert processed == ["first.jpg", "second.jpg"]
    assert [result.original_name for result in results] == ["first.jpg", "second.jpg"]
    assert results[0].status is RecognitionStatus.UNRECOGNIZED
    assert results[0].failure_reason == "PROCESS_FAILED: ocr crashed"
    assert results[1].status is RecognitionStatus.SUCCESS
    assert any("first.jpg" in message and "ocr crashed" in message for message in progress_messages)


def test_process_directory_keeps_processing_when_workbook_write_fails(tmp_path: Path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    first_path = input_dir / "first.jpg"
    second_path = input_dir / "second.jpg"
    first_path.write_bytes(b"first")
    second_path.write_bytes(b"second")
    progress_messages = []

    def fake_process_file(
        task: FileTask,
        _config: AppConfig,
        _ocr_engine: FakeOcrEngine,
        cancel_event=None,
    ) -> RecognitionResult:
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="HNKU6331795",
            elapsed_ms=1,
        )

    def fake_write_results(results, output_dir: Path, workbook_path: Path | None = None, comparison_report=None) -> Path:
        raise PermissionError("workbook is open")

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)
    monkeypatch.setattr(batch_module, "write_results", fake_write_results)

    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
        on_progress=progress_messages.append,
    )

    assert [result.original_name for result in results] == ["first.jpg", "second.jpg"]
    assert all(result.status is RecognitionStatus.SUCCESS for result in results)
    assert any("workbook is open" in message for message in progress_messages)


def test_process_directory_logs_expected_code_comparison(tmp_path: Path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    source_path = input_dir / "waybill.jpg"
    source_path.write_bytes(b"fake")
    progress_messages = []

    def fake_process_file(task, _config, _ocr_engine, cancel_event=None):
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="HNKU6331795",
            elapsed_ms=1,
        )

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)

    batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
        on_progress=progress_messages.append,
        expected_codes=["HNKU6331795", "GESU5903360"],
    )

    assert "箱号比对: 已匹配 1, 缺失 1, 多余 0, 格式无效 0" in progress_messages
    assert "缺失箱号: GESU5903360" in progress_messages
    workbook = load_workbook(output_dir / "识别结果.xlsx")
    assert "箱号比对" in workbook.sheetnames




def test_process_directory_skips_files_already_recorded_in_workbook(tmp_path: Path, monkeypatch):
    from waybill_ocr.output.excel_writer import write_results

    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    first_path = input_dir / "first.jpg"
    second_path = input_dir / "second.jpg"
    first_path.write_bytes(b"first")
    second_path.write_bytes(b"second")
    progress_messages = []
    processed = []
    existing_result = RecognitionResult(
        source_path=first_path,
        original_name=first_path.name,
        status=RecognitionStatus.SUCCESS,
        container_code="HNKU6331795",
        source=RecognitionSource.OCR,
        failure_reason=None,
        ocr_text="",
        elapsed_ms=1,
    )
    write_results([existing_result], output_dir)

    def fake_process_file(
        task: FileTask,
        _config: AppConfig,
        _ocr_engine: FakeOcrEngine,
        cancel_event=None,
    ) -> RecognitionResult:
        processed.append(task.relative_name)
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.SUCCESS,
            container_code="GESU5903360",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="GESU5903360",
            elapsed_ms=1,
        )

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)

    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
        on_progress=progress_messages.append,
    )

    assert processed == ["second.jpg"]
    assert [result.original_name for result in results] == ["first.jpg", "second.jpg"]
    workbook = load_workbook(output_dir / "识别结果.xlsx")
    assert [workbook.active["A2"].value, workbook.active["B2"].value] == ["first.jpg", "HNKU6331795"]
    assert [workbook.active["A3"].value, workbook.active["B3"].value] == ["second.jpg", "GESU5903360"]
    assert any("\u5df2\u8df3\u8fc7\u5df2\u5904\u7406\u6587\u4ef6: first.jpg" == message for message in progress_messages)
    assert any("first.jpg" in message and "HNKU6331795" in message for message in progress_messages)


def test_process_directory_does_not_save_evidence_image_for_non_success_result(tmp_path: Path, monkeypatch):
    from PIL import Image

    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    source_path = input_dir / "bad.jpg"
    Image.new("RGB", (32, 24), "white").save(source_path)

    def fake_process_file(
        task: FileTask,
        _config: AppConfig,
        _ocr_engine: FakeOcrEngine,
        cancel_event=None,
    ) -> RecognitionResult:
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.INVALID,
            container_code="HNKU6331794",
            source=RecognitionSource.OCR,
            failure_reason="INVALID_CHECK_DIGIT",
            ocr_text="HNKU6331794",
            elapsed_ms=1,
        )

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)

    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
    )

    evidence_dir = output_dir / "识别证据"
    assert not evidence_dir.exists()
    assert results[0].evidence_path is None
    workbook = load_workbook(output_dir / "识别结果.xlsx")
    headers = [cell.value for cell in workbook.active[1]]
    assert "证据截图" not in headers


def test_process_directory_skips_existing_results_by_relative_name_for_duplicate_names(tmp_path: Path, monkeypatch):
    from waybill_ocr.output.excel_writer import write_results

    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    first_dir = input_dir / "first"
    second_dir = input_dir / "second"
    first_dir.mkdir(parents=True)
    second_dir.mkdir(parents=True)
    first_path = first_dir / "waybill.pdf"
    second_path = second_dir / "waybill.pdf"
    first_path.write_bytes(b"first")
    second_path.write_bytes(b"second")
    processed = []
    existing_result = RecognitionResult(
        source_path=first_path,
        original_name=first_path.name,
        relative_name="first/waybill.pdf",
        status=RecognitionStatus.SUCCESS,
        container_code="HNKU6331795",
        source=RecognitionSource.OCR,
        failure_reason=None,
        ocr_text="",
        elapsed_ms=1,
    )
    write_results([existing_result], output_dir)

    def fake_process_file(task: FileTask, _config: AppConfig, _ocr_engine: FakeOcrEngine, cancel_event=None):
        processed.append(task.relative_name)
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            relative_name=task.relative_name,
            status=RecognitionStatus.SUCCESS,
            container_code="GESU5903360",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="GESU5903360",
            elapsed_ms=1,
        )

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)

    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
    )

    assert processed == ["second\\waybill.pdf"] or processed == ["second/waybill.pdf"]
    assert [result.relative_name for result in results] == ["first/waybill.pdf", processed[0]]
    assert [result.container_code for result in results] == ["HNKU6331795", "GESU5903360"]


def test_load_existing_results_falls_back_to_original_name_for_old_workbook(tmp_path: Path, monkeypatch):
    from waybill_ocr.output.excel_writer import write_results

    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    source_path = input_dir / "legacy.jpg"
    source_path.write_bytes(b"legacy")
    legacy_result = RecognitionResult(
        source_path=source_path,
        original_name=source_path.name,
        status=RecognitionStatus.SUCCESS,
        container_code="HNKU6331795",
        source=RecognitionSource.OCR,
        failure_reason=None,
        ocr_text="",
        elapsed_ms=1,
    )
    write_results([legacy_result], output_dir)
    workbook_path = output_dir / "识别结果.xlsx"
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    sheet.delete_cols(2)
    workbook.save(workbook_path)
    processed = []

    def fake_process_file(task: FileTask, _config: AppConfig, _ocr_engine: FakeOcrEngine, cancel_event=None):
        processed.append(task.relative_name)
        return legacy_result

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)

    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
    )

    assert processed == []
    assert [result.original_name for result in results] == ["legacy.jpg"]


def test_process_directory_uses_expected_list_review_code_before_copy(tmp_path: Path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    source_path = input_dir / "waybill.jpg"
    source_path.write_bytes(b"fake")

    def fake_process_file(task: FileTask, _config: AppConfig, _ocr_engine: FakeOcrEngine, cancel_event=None):
        return RecognitionResult(
            source_path=task.source_path,
            original_name=task.source_path.name,
            status=RecognitionStatus.UNRECOGNIZED,
            container_code=None,
            source=None,
            failure_reason="NO_CONTAINER_CANDIDATE",
            ocr_text="OCR HNKU6331795 GESU5903360",
            elapsed_ms=1,
        )

    monkeypatch.setattr(batch_module, "process_file", fake_process_file)

    results = batch_module.process_directory(
        input_dir=input_dir,
        output_dir=output_dir,
        config=AppConfig(),
        ocr_engine=FakeOcrEngine(),
        expected_codes=["GESU5903360"],
    )

    assert results[0].status is RecognitionStatus.UNRECOGNIZED
    assert results[0].review_code == "GESU5903360"
    assert any(path.name == "GESU5903360-待确认.jpg" for path in output_dir.rglob("*.jpg"))
    workbook = load_workbook(output_dir / "识别结果.xlsx")
    assert workbook.active["B2"].value == "GESU5903360"
    assert workbook.active["C2"].value == RecognitionStatus.UNRECOGNIZED.value