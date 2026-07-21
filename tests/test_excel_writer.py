from pathlib import Path

from openpyxl import load_workbook

from waybill_ocr.constants import RESULT_WORKBOOK_NAME
from waybill_ocr.models import RecognitionResult, RecognitionSource, RecognitionStatus
from waybill_ocr.output import excel_writer
from waybill_ocr.output.excel_writer import write_results


VISIBLE_HEADERS = ["原始文件名", "识别箱号", "识别状态", "处理耗时ms", "备注"]
INTERNAL_HEADERS = [
    "原始文件名",
    "相对路径",
    "原始识别箱号",
    "识别来源",
    "失败原因",
    "复核候选",
    "识别状态",
    "输出相对路径",
]

def test_write_results_creates_workbook_with_clean_recognition_rows(tmp_path: Path):
    source = tmp_path / "HNKU6331795.jpg"
    result = RecognitionResult(
        source_path=source,
        original_name=source.name,
        status=RecognitionStatus.SUCCESS,
        container_code="HNKU6331795",
        source=RecognitionSource.OCR,
        failure_reason=None,
        ocr_text="HNKU6331795",
        elapsed_ms=123,
    )

    workbook_path = write_results([result], tmp_path)

    assert workbook_path == tmp_path / RESULT_WORKBOOK_NAME
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    assert sheet.title == "识别结果"
    assert [cell.value for cell in sheet[1]] == VISIBLE_HEADERS
    assert [cell.value for cell in sheet[2]] == [
        "HNKU6331795.jpg",
        "HNKU6331795",
        "正确识别",
        123,
        None,
    ]


def test_write_results_highlights_non_success_rows(tmp_path: Path):
    success_source = tmp_path / "success.jpg"
    invalid_source = tmp_path / "invalid.jpg"
    unrecognized_source = tmp_path / "unrecognized.jpg"
    results = [
        RecognitionResult(
            source_path=success_source,
            original_name=success_source.name,
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="HNKU6331795",
            elapsed_ms=1,
        ),
        RecognitionResult(
            source_path=invalid_source,
            original_name=invalid_source.name,
            status=RecognitionStatus.INVALID,
            container_code="HNKU6331794",
            source=RecognitionSource.OCR,
            failure_reason="INVALID_CHECK_DIGIT",
            ocr_text="HNKU6331794",
            elapsed_ms=2,
        ),
        RecognitionResult(
            source_path=unrecognized_source,
            original_name=unrecognized_source.name,
            status=RecognitionStatus.UNRECOGNIZED,
            container_code=None,
            source=None,
            failure_reason="NO_CONTAINER_CODE",
            ocr_text="",
            elapsed_ms=3,
        ),
    ]

    workbook_path = write_results(results, tmp_path)

    sheet = load_workbook(workbook_path).active
    assert all(cell.fill.fill_type is None for cell in sheet[2])
    for row_number in (3, 4):
        assert all(cell.fill.fill_type == "solid" for cell in sheet[row_number][: len(VISIBLE_HEADERS)])
        assert all(cell.fill.fgColor.rgb in {"00FFC7CE", "FFFFC7CE"} for cell in sheet[row_number][: len(VISIBLE_HEADERS)])


def test_write_results_outputs_review_note_without_source_column(tmp_path: Path):
    source = tmp_path / "waybill.jpg"
    result = RecognitionResult(
        source_path=source,
        original_name=source.name,
        status=RecognitionStatus.SUCCESS,
        container_code="HNKU6331795",
        source=RecognitionSource.OCR_REPAIRED,
        failure_reason=None,
        ocr_text="OCR HINKU6331795",
        elapsed_ms=123,
        review_note="OCR修正原始片段: HINKU6331795",
    )

    workbook_path = write_results([result], tmp_path)

    sheet = load_workbook(workbook_path).active
    headers = [cell.value for cell in sheet[1]]
    assert "识别来源" not in headers
    assert sheet["E1"].value == "备注"
    assert sheet["E2"].value == "OCR修正原始片段: HINKU6331795"


def test_write_results_displays_review_code_for_non_success_result(tmp_path: Path):
    source = tmp_path / "waybill.pdf"
    result = RecognitionResult(
        source_path=source,
        original_name=source.name,
        status=RecognitionStatus.INVALID,
        container_code="YYCU6002610",
        source=RecognitionSource.OCR,
        failure_reason="INVALID_CHECK_DIGIT",
        ocr_text="YYCU6002610",
        elapsed_ms=123,
        review_note="疑似修正: YYCU6002610 -> YYCU6003610",
        review_code="YYCU6003610",
    )

    workbook_path = write_results([result], tmp_path)

    sheet = load_workbook(workbook_path).active
    assert sheet["B2"].value == "YYCU6003610"
    assert sheet["C2"].value == "箱号错误"
    assert "YYCU6002610 -> YYCU6003610" in sheet["E2"].value
    assert "校验位不正确" in sheet["E2"].value


def test_write_results_sets_original_name_column_width(tmp_path: Path):
    source = tmp_path / "long-original-file-name.pdf"
    result = RecognitionResult(
        source_path=source,
        original_name=source.name,
        status=RecognitionStatus.SUCCESS,
        container_code="HNKU6331795",
        source=RecognitionSource.OCR,
        failure_reason=None,
        ocr_text="HNKU6331795",
        elapsed_ms=1,
    )

    workbook_path = write_results([result], tmp_path)

    sheet = load_workbook(workbook_path).active
    assert sheet.column_dimensions["A"].width >= 48


def test_write_results_adds_comparison_sheet(tmp_path: Path):
    from waybill_ocr.container_code.expected_codes import compare_expected_codes

    source = tmp_path / "waybill.pdf"
    result = RecognitionResult(
        source_path=source,
        original_name=source.name,
        status=RecognitionStatus.SUCCESS,
        container_code="HNKU6331795",
        source=RecognitionSource.OCR,
        failure_reason=None,
        ocr_text="HNKU6331795",
        elapsed_ms=1,
    )
    report = compare_expected_codes(
        expected_codes=["HNKU6331795", "GESU5903360"],
        results=[result],
        invalid_expected_entries=["BAD-CODE"],
    )

    workbook_path = write_results([result], tmp_path, comparison_report=report)

    workbook = load_workbook(workbook_path)
    sheet = workbook["箱号比对"]
    assert [cell.value for cell in sheet[1]] == [
        "预期箱号",
        "状态",
        "对应文件",
        "多余识别箱号",
        "格式无效",
    ]
    assert [cell.value for cell in sheet[2][:5]] == [
        "HNKU6331795",
        "已识别",
        "waybill.pdf",
        None,
        "BAD-CODE",
    ]
    assert [cell.value for cell in sheet[3][:3]] == ["GESU5903360", "缺失", None]

def test_write_results_omits_evidence_and_diagnostic_columns_from_visible_sheet(tmp_path: Path):
    source = tmp_path / "waybill.jpg"
    evidence = tmp_path / "识别证据" / "waybill.png"
    result = RecognitionResult(
        source_path=source,
        original_name=source.name,
        status=RecognitionStatus.UNRECOGNIZED,
        container_code=None,
        source=None,
        failure_reason="NO_CONTAINER_CANDIDATE",
        ocr_text="",
        elapsed_ms=1,
        evidence_path=evidence,
        relative_name="nested/waybill.jpg",
    )

    workbook_path = write_results([result], tmp_path)

    sheet = load_workbook(workbook_path).active
    headers = [cell.value for cell in sheet[1]]
    assert "证据截图" not in headers
    assert "识别来源" not in headers
    assert "失败原因" not in headers
    assert "相对路径" not in headers
    assert headers == VISIBLE_HEADERS


def test_write_results_stores_diagnostics_in_hidden_internal_index(tmp_path: Path):
    source = tmp_path / "nested" / "waybill.pdf"
    result = RecognitionResult(
        source_path=source,
        original_name=source.name,
        relative_name="nested/waybill.pdf",
        status=RecognitionStatus.INVALID,
        container_code="YYCU6002610",
        source=RecognitionSource.OCR,
        failure_reason="INVALID_CHECK_DIGIT",
        ocr_text="YYCU6002610",
        elapsed_ms=1,
        review_code="YYCU6003610",
        output_relative_path="箱号错误/YYCU6003610-待确认.pdf",
    )

    workbook_path = write_results([result], tmp_path)

    workbook = load_workbook(workbook_path)
    index_sheet = workbook["内部索引"]
    assert index_sheet.sheet_state == "hidden"
    assert [cell.value for cell in index_sheet[1]] == INTERNAL_HEADERS
    assert [cell.value for cell in index_sheet[2]] == [
        "waybill.pdf",
        "nested/waybill.pdf",
        "YYCU6002610",
        "OCR",
        "INVALID_CHECK_DIGIT",
        "YYCU6003610",
        "箱号错误",
        "箱号错误/YYCU6003610-待确认.pdf",
    ]


def test_write_results_replaces_workbook_atomically(tmp_path: Path, monkeypatch):
    source = tmp_path / "waybill.pdf"
    result = RecognitionResult(
        source_path=source,
        original_name=source.name,
        status=RecognitionStatus.SUCCESS,
        container_code="HNKU6331795",
        source=RecognitionSource.OCR,
        failure_reason=None,
        ocr_text="HNKU6331795",
        elapsed_ms=1,
    )
    replace_calls: list[tuple[Path, Path]] = []
    real_replace = excel_writer.os.replace

    def recording_replace(source_path, target_path):
        replace_calls.append((Path(source_path), Path(target_path)))
        real_replace(source_path, target_path)

    monkeypatch.setattr(excel_writer.os, "replace", recording_replace)

    workbook_path = write_results([result], tmp_path)

    assert workbook_path.exists()
    assert len(replace_calls) == 1
    assert replace_calls[0][1] == workbook_path
    assert replace_calls[0][0] != workbook_path
    assert not list(tmp_path.glob("*.tmp.xlsx"))

def test_write_results_comparison_sheet_uses_expected_order_and_status(tmp_path: Path):
    from waybill_ocr.container_code.expected_codes import compare_expected_codes

    results = [
        RecognitionResult(
            source_path=tmp_path / "one.pdf",
            original_name="one.pdf",
            status=RecognitionStatus.SUCCESS,
            container_code="HNKU6331795",
            source=RecognitionSource.OCR,
            failure_reason=None,
            ocr_text="",
            elapsed_ms=1,
        ),
        RecognitionResult(
            source_path=tmp_path / "two.pdf",
            original_name="two.pdf",
            status=RecognitionStatus.UNRECOGNIZED,
            container_code=None,
            source=None,
            failure_reason="NO_CONTAINER_CANDIDATE",
            ocr_text="GESU5903360",
            elapsed_ms=1,
            review_code="GESU5903360",
        ),
    ]
    report = compare_expected_codes(["GESU5903360", "HNKU6331795", "MSCU1234566"], results)

    workbook_path = write_results(results, tmp_path, comparison_report=report)

    workbook = load_workbook(workbook_path)
    sheet = workbook["箱号比对"]
    assert [cell.value for cell in sheet[1]] == ["预期箱号", "状态", "对应文件", "多余识别箱号", "格式无效"]
    assert [cell.value for cell in sheet[2][:3]] == ["GESU5903360", "待确认命中", "two.pdf"]
    assert [cell.value for cell in sheet[3][:3]] == ["HNKU6331795", "已识别", "one.pdf"]
    assert [cell.value for cell in sheet[4][:3]] == ["MSCU1234566", "缺失", None]
    assert (tmp_path / "缺失箱号清单.txt").read_text(encoding="utf-8") == "GESU5903360\nMSCU1234566\n"


def test_write_results_missing_codes_file_says_none_when_no_missing(tmp_path: Path):
    from waybill_ocr.container_code.expected_codes import compare_expected_codes

    result = RecognitionResult(
        source_path=tmp_path / "one.pdf",
        original_name="one.pdf",
        status=RecognitionStatus.SUCCESS,
        container_code="HNKU6331795",
        source=RecognitionSource.OCR,
        failure_reason=None,
        ocr_text="",
        elapsed_ms=1,
    )
    report = compare_expected_codes(["HNKU6331795"], [result])

    write_results([result], tmp_path, comparison_report=report)

    assert (tmp_path / "缺失箱号清单.txt").read_text(encoding="utf-8") == "无缺失箱号\n"
